#!/usr/bin/env python3
"""Harvest guild quest catalogs from their real sources.

This is the configurator seam of the absorption engine: sources.json says which guild,
which source, which adapter; each adapter's only contract is "emit quests that conform
to schema.md". Content passes through verbatim; anything odd lands in the anomalies
report for the guild to rule on — never silently fixed.

Usage:
  python harvest.py                     # harvest every enabled source in sources.json
  python harvest.py slayers-summons     # harvest one source by id

Outputs, per source:
  <output>.json            the canonical quest catalog
  <output>-anomalies.md    everything the guild should look at

The sheet-xlsx adapter needs openpyxl (pip install openpyxl). Everything else is
standard library only.
"""
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))

SLOT_PATTERN = re.compile(r"(\w+):(?=\s|$)")
IMAGE_SLOT = re.compile(r"^image\d*$")
# the parameter that names which quest/badge the bot credits, and its value
NAME_PARAM = re.compile(r"(?:summons_type|badge_name):\s*(.*?)(?=\s+\w+:|$)")


# ---------------------------------------------------------------- shared helpers

def slugify(name):
    slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
    return slug


def anom(message, kind, row=None, tab=None, quest_id=None, quest_name=None):
    """One structured anomaly. The prose stays verbatim in `message`; the prefix that
    used to be baked into the string (row N / **Name**) is reconstructed from the
    structure by write_anomalies, and joined to rows/quests by the provenance view."""
    return {
        "kind": kind,
        "row": row,
        "tab": tab,
        "quest_id": quest_id,
        "quest_name": quest_name,
        "message": message,
    }


def col_letter(index):
    """0-based column index -> spreadsheet letter (A, B, ... Z, AA, ...)."""
    letters = ""
    index += 1
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(ord("A") + rem) + letters
    return letters


class Provenance:
    """Records, per source, what the adapter actually did: which columns it read and
    what canonical fields they became, and the fate of every row it saw — with the
    verbatim cell values, so a leader can hold the page next to their own sheet.
    Emitted as a sidecar (<output>-provenance.json); never flows into the catalog."""

    CELL_LIMIT = 500

    def __init__(self, source, mode="sheet"):
        self.source = {k: source.get(k) for k in
                       ("id", "guild", "era", "adapter", "path", "tab", "url", "retrieved")}
        self.mode = mode
        self.tabs = []
        self._cur = None

    def tab(self, name, columns, header_row=None):
        """Start a tab scope. columns = [(index, header, fields, note), ...]."""
        self._cur = {
            "tab": name,
            "header_row": header_row,
            "columns": [
                {"index": i, "letter": col_letter(i), "header": h, "fields": f, "note": n}
                for (i, h, f, n) in columns
            ],
            "rows": [],
        }
        self.tabs.append(self._cur)

    def row(self, i, values, outcome, quest_id=None, reason=None, **extra):
        """Record one source row. `i` is the 1-based spreadsheet row. Cell echoes are
        limited to mapped columns (column A only for banner/section rows) and
        truncated so the sidecar stays bounded."""
        entry = {"row": i, "outcome": outcome}
        if quest_id is not None:
            entry["quest_id"] = quest_id
        if reason is not None:
            entry["reason"] = reason
        entry.update(extra)
        if values is not None and outcome != "blank" and self._cur is not None:
            columns = self._cur["columns"]
            if outcome in ("banner", "section", "filler"):
                columns = columns[:1]
            cells = {}
            truncated = False
            for c in columns:
                idx = c["index"]
                v = values[idx] if idx < len(values) else None
                if v is None or str(v).strip() == "":
                    continue
                s = str(v)
                if len(s) > self.CELL_LIMIT:
                    s = s[: self.CELL_LIMIT] + "…"
                    truncated = True
                cells[c["letter"]] = s
            if cells:
                entry["cells"] = cells
            if truncated:
                entry["truncated"] = True
        self._cur["rows"].append(entry)

    def finish(self, anomalies, quest_count):
        def tally(outcome):
            return sum(1 for t in self.tabs for r in t["rows"] if r["outcome"] == outcome)

        return {
            "schema_version": 1,
            "mode": self.mode,
            "source": self.source,
            "tabs": self.tabs,
            "anomalies": anomalies,
            "counts": {
                "rows_seen": sum(len(t["rows"]) for t in self.tabs),
                "quests": quest_count,
                "skipped": tally("skipped"),
                "blank": tally("blank"),
                "anomalies": len(anomalies),
            },
        }


def parse_evidence(bot_command, requirements_text):
    """Derive the evidence spec from the command's slots (machine truth) and the
    requirements text's emoji grammar (human truth). Mismatches are anomalies."""
    cmd = bot_command or ""
    slots = SLOT_PATTERN.findall(cmd)
    # the name parameter (summons_type:/badge_name:) is the selector, not evidence;
    # participants:/url slots may carry instruction text instead of being empty,
    # so those are detected by substring, not by the empty-slot pattern
    evidence_slots = [s for s in slots if s not in ("summons_type", "badge_name")]
    return {
        "screenshots": sum(1 for s in evidence_slots if IMAGE_SLOT.match(s)),
        "video_alternative": "\U0001F39E" in (requirements_text or ""),  # 🎞️
        "link": "summons_url:" in cmd,
        "group_turnin": "participants:" in cmd,
        "notes": "summons_notes:" in cmd or "badge_notes:" in cmd,
    }


def cross_check(quest, anomalies):
    """Compare emoji grammar in the text against the command slots."""
    text = quest["requirements"] or ""
    ev = quest["evidence"]
    camera_count = text.count("\U0001F4F8")  # 📸

    def flag(message):
        anomalies.append(anom(message, "evidence_mismatch",
                              quest_id=quest["quest_id"], quest_name=quest["name"]))

    if quest["auto_checked"]:
        return
    if camera_count and ev["screenshots"] and camera_count != ev["screenshots"]:
        flag(
            f"requirements text shows {camera_count} camera emoji "
            f"but the bot command has {ev['screenshots']} image slot(s). "
            f"Which is right?"
        )
    if "\U0001F517" in text and not ev["link"]:  # 🔗
        flag(
            f"requirements mention a 🔗 link but the bot command "
            f"has no summons_url: slot."
        )
    if "\U0001F91C" in text and not ev["group_turnin"]:  # 🤜
        flag(
            f"requirements mention 🤜🤛 group turn-in but the bot "
            f"command has no participants: slot."
        )
    note = quest.get("evidence_note") or ""
    if ev["screenshots"] == 0 and re.search(r"screenshot|photo|image", note, re.I):
        flag(
            f"the evidence note asks for a screenshot/photo "
            f"({note[:60]!r}) but the bot command has no image slot."
        )


def finish_catalog(source, quests, anomalies):
    """Shared post-pass: ids unique, commands unique, wrap in the catalog envelope."""
    seen_ids = {}
    seen_commands = {}
    for q in quests:
        if q["quest_id"] in seen_ids:
            anomalies.append(anom(
                f"quest_id `{q['quest_id']}` collides with "
                f"**{seen_ids[q['quest_id']]}** — one of them needs a distinct name.",
                "duplicate_id", quest_id=q["quest_id"], quest_name=q["name"],
            ))
        seen_ids[q["quest_id"]] = q["name"]

        cmd = q["bot_command"]
        if cmd:
            m = NAME_PARAM.search(cmd)
            credited = m.group(1).strip() if m else None
            if credited is None:
                anomalies.append(anom(
                    f"bot command has no summons_type:/badge_name: "
                    f"parameter — the bot cannot tell what is being turned in: `{cmd}`",
                    "missing_name_param", quest_id=q["quest_id"], quest_name=q["name"],
                ))
            else:
                if credited in seen_commands and seen_commands[credited] != q["name"]:
                    anomalies.append(anom(
                        f"bot command credits `{credited}` which is "
                        f"also credited by **{seen_commands[credited]}** — likely a "
                        f"copy-paste slip in the source. The bot would credit the "
                        f"wrong quest.",
                        "duplicate_credit", quest_id=q["quest_id"], quest_name=q["name"],
                    ))
                else:
                    seen_commands[credited] = q["name"]
                if credited.lower() != q["name"].lower():
                    anomalies.append(anom(
                        f"bot command credits `{credited}` — not the "
                        f"quest's own name. Typo or intentional?",
                        "credit_mismatch", quest_id=q["quest_id"], quest_name=q["name"],
                    ))
        cross_check(q, anomalies)

    return {
        "schema_version": 1,
        "guild": source["guild"],
        "era": source["era"],
        "source": {
            "kind": source["adapter"],
            "detail": f"{source.get('path')} :: {source.get('tab', '')}".strip(" :"),
            "url": source.get("url"),
            "retrieved": source.get("retrieved"),
        },
        "quests": quests,
    }


# ---------------------------------------------------------------- adapters

# which columns the sheet-xlsx adapter reads, and what they become
SHEET_COLUMNS = [
    (0, "Name", ["name", "quest_id"], "quest_id is a slug of the name"),
    (1, "Coopable?", ["coopable"], None),
    (2, "Category", ["category"], None),
    (3, "Turn-in Requirements", ["requirements", "evidence.video_alternative"],
     "verbatim; emoji grammar cross-checked against the bot command"),
    (4, "Bot Template", ["bot_command", "auto_checked", "evidence"],
     "the command's slots become the evidence spec"),
]


def adapt_sheet_xlsx(source):
    """Harvest a normalized quest tab from a guild tracker workbook.
    Expects columns: Name | Coopable? | Category | Turn-in Requirements | Bot Template."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("The sheet-xlsx adapter needs openpyxl: pip install openpyxl")

    path = os.path.normpath(os.path.join(HERE, source["path"]))
    wb = load_workbook(path, read_only=True)
    if source["tab"] not in wb.sheetnames:
        raise SystemExit(f"tab {source['tab']!r} not found in {path}")
    ws = wb[source["tab"]]

    prov = Provenance(source)
    prov.tab(source["tab"], SHEET_COLUMNS, header_row=1)

    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    expected = ["name", "coopable?", "category", "turn-in requirements", "bot template"]
    anomalies = []
    prov.row(1, rows[0], "header")
    if header[: len(expected)] != expected:
        anomalies.append(anom(
            f"header row is {header[:5]} — expected {expected}. "
            f"Columns may have moved; harvest read them positionally.",
            "header_mismatch", row=1,
        ))

    quests = []
    skipped = 0
    for i, row in enumerate(rows[1:], start=2):
        name = (str(row[0]).strip() if row[0] is not None else "")
        if not name:
            prov.row(i, row, "blank")
            continue
        coopable, category, req, cmd = row[1], row[2], row[3], row[4]
        if category is None and req is None and cmd is None:
            anomalies.append(anom(
                "has a name but no other data — skipped.",
                "skipped_row", row=i, quest_name=name,
            ))
            prov.row(i, row, "skipped", reason="has a name but no other data")
            skipped += 1
            continue
        req = str(req).strip() if req is not None else ""
        cmd = str(cmd).strip() if cmd is not None else None
        # the sheet marks meta-quests by putting "No submission, auto-checked" in the
        # template column instead of a command
        auto = (
            cmd is None
            or not cmd.startswith("/")
            or "auto-checked" in req.lower()
        )
        if cmd is not None and not cmd.startswith("/") and "auto" not in cmd.lower():
            anomalies.append(anom(
                f"bot template is not a command and does not say "
                f"auto-checked: {cmd!r}. Treated as auto-checked — is that right?",
                "not_a_command", row=i, quest_name=name,
            ))
        quest = {
            "quest_id": slugify(name),
            "name": name,
            "category": str(category).strip() if category is not None else "",
            "coopable": bool(coopable),
            "requirements": req,
            "reward": None,
            "evidence": parse_evidence(None if auto else cmd, req),
            "evidence_note": None,
            "bot_command": None if auto else cmd,
            "auto_checked": auto,
            "venue": "in_game",
            "trigger": None,
        }
        quests.append(quest)
        prov.row(i, row, "quest", quest_id=quest["quest_id"])

    if skipped:
        anomalies.append(anom(
            f"{skipped} row(s) skipped for missing data (listed above).", "skip_summary"))
    return finish_catalog(source, quests, anomalies), anomalies, prov


def adapt_gm_template(source):
    """A GM hands us a filled template that is already in catalog shape: validate the
    envelope, regenerate ids, run the same cross-checks. Content passes through."""
    path = os.path.normpath(os.path.join(HERE, source["path"]))
    with open(path, encoding="utf-8-sig") as f:
        data = json.load(f)
    anomalies = []
    quests = []
    prov = Provenance(source, mode="passthrough")
    prov.tab(None, [])
    for i, q in enumerate(data.get("quests", []), start=1):
        name = (q.get("name") or "").strip()
        if not name:
            anomalies.append(anom(
                f"a quest entry has no name — skipped: {json.dumps(q)[:80]}",
                "no_name", row=i,
            ))
            prov.row(i, None, "skipped", reason="quest entry has no name")
            continue
        cmd = q.get("bot_command")
        auto = bool(q.get("auto_checked")) or cmd is None
        prov.row(i, None, "quest", quest_id=q.get("quest_id") or slugify(name))
        quests.append({
            "quest_id": q.get("quest_id") or slugify(name),
            "name": name,
            "category": q.get("category", ""),
            "coopable": bool(q.get("coopable", False)),
            "requirements": q.get("requirements", ""),
            "evidence": q.get("evidence") or parse_evidence(None if auto else cmd, q.get("requirements", "")),
            "bot_command": None if auto else cmd,
            "auto_checked": auto,
            "venue": q.get("venue", "in_game"),
            "trigger": q.get("trigger"),
        })
    return finish_catalog(source, quests, anomalies), anomalies, prov


# which columns the ranger-xlsx adapter reads, per tab
RANGER_BADGES_COLUMNS = [
    (0, "Section title / shared marker", ["category", "coopable"],
     "section title rows set the category; 'shared' marks group badges"),
    (1, "Badge", ["name", "quest_id", "name_note"],
     "a parenthesized second line moves to name_note"),
    (3, "Description", ["requirements"], None),
    (7, "Turn-in", ["bot_command", "evidence"], None),
    (10, "Required Screenshots", ["evidence_note"], None),
]
RANGER_QUESTS_COLUMNS = [
    (0, "Quest (name + reward)", ["name", "quest_id", "reward"],
     "reward parsed out of the same cell"),
    (1, "Description", ["requirements"], None),
    (3, "Turn-in", ["bot_command", "evidence"], None),
]


def adapt_ranger_xlsx(source):
    """Harvest the Ranger tracker: a sectioned 'Badges' tab (category title rows, a
    'shared' group marker, per-badge evidence notes, an IRL section) plus a narrative
    'Quests' tab (name and reward folded into one cell)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("The ranger-xlsx adapter needs openpyxl: pip install openpyxl")

    path = os.path.normpath(os.path.join(HERE, source["path"]))
    wb = load_workbook(path, read_only=True)
    anomalies = []
    quests = []
    prov = Provenance(source)

    # --- Badges tab: repeated sections, columns fixed at
    #     shared-marker=0, name=1, description=3, turn-in=7, evidence-note=10
    section = re.compile(r"^(.*\bBadges)\s*$")
    ws = wb["Badges"]
    prov.tab("Badges", RANGER_BADGES_COLUMNS)
    category = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cell0 = str(row[0]).strip() if row[0] is not None else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        m = section.match(cell0)
        if m and not name:
            category = re.sub(r"\s+", " ", m.group(1))
            prov.row(i, row, "section", category=category)
            continue
        if not name or name == "Badge":
            prov.row(i, row, "header" if name == "Badge" else "blank")
            continue
        # names sometimes carry a parenthesized annotation on a second line,
        # e.g. "Igloo\n(Ranger Station)" — the bot credits the bare name
        name_note = None
        name_lines = [ln.strip() for ln in name.splitlines() if ln.strip()]
        if len(name_lines) > 1 and all(ln.startswith("(") for ln in name_lines[1:]):
            name = name_lines[0]
            name_note = " ".join(name_lines[1:])
        if category is None:
            anomalies.append(anom(
                "appears before any section title — skipped.",
                "no_section", row=i, tab="Badges", quest_name=name,
            ))
            prov.row(i, row, "skipped", reason="appears before any section title")
            continue
        desc = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        cmd = str(row[7]).strip() if len(row) > 7 and row[7] is not None else None
        note = str(row[10]).strip() if len(row) > 10 and row[10] is not None else None
        if not cmd:
            anomalies.append(anom(
                "no turn-in command — skipped. Auto-checked, retired, or a slip?",
                "no_command", row=i, tab="Badges", quest_name=name,
            ))
            prov.row(i, row, "skipped", reason="no turn-in command")
            continue
        prov.row(i, row, "quest", quest_id=slugify(name))
        quests.append({
            "quest_id": slugify(name),
            "name": name,
            "name_note": name_note,
            "category": category,
            "coopable": "shared" in cell0.lower(),
            "requirements": desc,
            "reward": None,
            "evidence": parse_evidence(cmd, f"{desc}\n{note or ''}"),
            "evidence_note": note,
            "bot_command": cmd,
            "auto_checked": False,
            "venue": "irl" if category.lower().startswith("irl") else "in_game",
            "trigger": None,
        })

    # --- Quests tab: name + reward folded into col 0, description col 1, turn-in col 3
    ws = wb["Quests"]
    prov.tab("Quests", RANGER_QUESTS_COLUMNS)
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        blob = str(row[0]).strip() if row[0] is not None else ""
        cmd = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
        if not blob or blob == "Quest" or not cmd:
            outcome = "blank" if not blob else ("header" if blob == "Quest" else "filler")
            prov.row(i, row, outcome)
            continue
        lines = [ln.strip() for ln in blob.splitlines() if ln.strip()]
        name = lines[0] if lines else ""
        reward = None
        for j, ln in enumerate(lines):
            if ln.lower().startswith("reward"):
                inline = ln.split(":", 1)[1].strip() if ":" in ln else ""
                reward = inline or " ".join(lines[j + 1:]) or None
                break
        if not name:
            anomalies.append(anom(
                f"has a turn-in command but no readable name — skipped: {blob[:60]!r}",
                "unreadable_name", row=i, tab="Quests",
            ))
            prov.row(i, row, "skipped", reason="turn-in command but no readable name")
            continue
        desc = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        prov.row(i, row, "quest", quest_id=slugify(name))
        quests.append({
            "quest_id": slugify(name),
            "name": name,
            "category": "Quests",
            "coopable": True,  # tab header: collaboration allowed and encouraged
            "requirements": desc,
            "reward": reward,
            "evidence": parse_evidence(cmd, desc),
            "evidence_note": None,
            "bot_command": cmd,
            "auto_checked": False,
            "venue": "in_game",
            "trigger": None,
        })

    return finish_catalog(source, quests, anomalies), anomalies, prov


# which columns the creator-events adapter reads, and what they become
CREATOR_COLUMNS = (
    [
        (0, "Event Name", ["name", "quest_id"], None),
        (1, "Creator", ["event.creator"], None),
        (2, "Status", ["event.status"], None),
        (3, "Run time (minutes)", ["event.run_time_minutes"], None),
        (4, "Type", ["category"], None),
        (5, "Biome", ["event.biome"], None),
        (6, "Special restrictions", ["requirements"], None),
        (7, "Players", ["coopable", "event.players"], "coopable when the max count > 1"),
        (9, "Gear Level", ["event.gear_level"], None),
    ]
    + [(c, f"Unique {c - 10}", ["reward", "event.uniques"], None) for c in range(11, 19)]
    + [
        (19, "Blue Mushies", ["event.blue_mushrooms"], None),
        (20, "Total Gold", ["event.total_gold"], None),
        (21, "Gems / rare valuables", ["event.other_loot"], None),
        (22, "Mats in containers", ["event.other_loot"], None),
        (23, "Choppables / breakables", ["event.other_loot"], None),
        (24, "Cape Set", ["event.cape_set"], None),
    ]
)


def adapt_creator_events_xlsx(source):
    """Harvest a creator events tracker: one 'Events (E<n>)' tab per era. Events have
    no bot turn-in (a GM runs them live), so every entry is auto_checked; the event's
    operational facts (creator, status, biome, loadout uniques, loot) ride along
    verbatim in an `event` object."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("The creator-events-xlsx adapter needs openpyxl: pip install openpyxl")

    path = os.path.normpath(os.path.join(HERE, source["path"]))
    wb = load_workbook(path, read_only=True, data_only=True)
    if source["tab"] not in wb.sheetnames:
        raise SystemExit(f"tab {source['tab']!r} not found in {path}")
    ws = wb[source["tab"]]
    ws.reset_dimensions()

    def text(row, i):
        return str(row[i]).strip() if len(row) > i and row[i] is not None else ""

    rows = list(ws.iter_rows(values_only=True))
    anomalies = []
    prov = Provenance(source)

    # the tab has a banner row; the real header is the row whose first cell ends in
    # "Event Name". Columns are fixed positionally from there.
    header_idx = None
    for i, row in enumerate(rows):
        if text(row, 0).lower().endswith("event name"):
            header_idx = i
            break
    if header_idx is None:
        raise SystemExit(f"no 'Event Name' header row found in tab {source['tab']!r}")
    prov.tab(source["tab"], CREATOR_COLUMNS, header_row=header_idx + 1)
    for i, row in enumerate(rows[:header_idx], start=1):
        prov.row(i, row, "banner")
    prov.row(header_idx + 1, rows[header_idx], "header")
    header = [text(rows[header_idx], i).lower() for i in range(10)]
    expected = {1: "creator", 2: "status", 4: "type", 5: "biome",
                6: "special restrictions", 7: "players", 9: "gear level"}
    for col, want in expected.items():
        if want not in header[col]:
            anomalies.append(anom(
                f"header column {col} is {header[col]!r} — expected it to contain "
                f"{want!r}. Columns may have moved; harvest read them positionally.",
                "header_mismatch", row=header_idx + 1,
            ))

    quests = []
    skipped = 0
    for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        name = text(row, 0)
        if not name:
            prov.row(i, row, "blank")
            continue
        creator, status, event_type = text(row, 1), text(row, 2), text(row, 4)
        uniques = [text(row, c) for c in range(11, 19) if text(row, c)]
        if not creator and not status and not event_type and not uniques:
            anomalies.append(anom(
                "has a name but no other data — skipped.",
                "skipped_row", row=i, quest_name=name,
            ))
            prov.row(i, row, "skipped", reason="has a name but no other data")
            skipped += 1
            continue

        players = text(row, 7)
        player_counts = [int(n) for n in re.findall(r"\d+", players)]
        if not player_counts:
            anomalies.append(anom(
                f"players is {players!r} — could not read a count. "
                f"Treated as solo (not coopable).",
                "unreadable_players", row=i, quest_name=name,
            ))
        coopable = bool(player_counts) and max(player_counts) > 1

        if status.lower() == "tbd":
            anomalies.append(anom(
                "status is TBD — is this event ready to run?",
                "tbd_status", quest_id=slugify(name), quest_name=name,
            ))

        restrictions = text(row, 6)
        loot = [t for t in (text(row, 21), text(row, 22), text(row, 23)) if t]
        prov.row(i, row, "quest", quest_id=slugify(name))
        quests.append({
            "quest_id": slugify(name),
            "name": name,
            "category": event_type,
            "coopable": coopable,
            "requirements": restrictions,
            "reward": ", ".join(uniques) or None,
            "evidence": parse_evidence(None, restrictions),
            "evidence_note": None,
            "bot_command": None,
            "auto_checked": True,
            "venue": "in_game",
            "trigger": None,
            "event": {
                "creator": creator,
                "status": status,
                "run_time_minutes": text(row, 3),
                "biome": text(row, 5),
                "players": players,
                "gear_level": text(row, 9),
                "uniques": uniques,
                "blue_mushrooms": text(row, 19),
                "total_gold": text(row, 20),
                "other_loot": loot,
                "cape_set": text(row, 24),
            },
        })

    if skipped:
        anomalies.append(anom(
            f"{skipped} row(s) skipped for missing data (listed above).", "skip_summary"))
    return finish_catalog(source, quests, anomalies), anomalies, prov


def adapt_discord_export(source):
    """Reserved: harvest quests straight from a Discord channel export. This is where
    the absorption engine plugs in. Deliberately unimplemented."""
    raise SystemExit(
        "The discord-export adapter is a stub — it marks the seam where the absorption "
        "engine will plug in. Use sheet-xlsx or gm-template for now."
    )


ADAPTERS = {
    "sheet-xlsx": adapt_sheet_xlsx,
    "ranger-xlsx": adapt_ranger_xlsx,
    "creator-events-xlsx": adapt_creator_events_xlsx,
    "gm-template": adapt_gm_template,
    "discord-export": adapt_discord_export,
}


# ---------------------------------------------------------------- driver

def format_anomaly(a):
    """Reconstruct the human-facing line from a structured anomaly: the row/quest
    prefix that used to be baked into the prose."""
    if a["row"] is not None:
        label = f"row {a['row']}"
        if a["tab"]:
            label = f"{a['tab']} {label}"
        if a["quest_name"]:
            label += f" ({a['quest_name']!r})"
        return f"{label}: {a['message']}"
    if a["quest_name"]:
        return f"**{a['quest_name']}**: {a['message']}"
    return a["message"]


def write_anomalies(path, source, anomalies):
    lines = [
        f"# Anomalies — {source['guild']} quest catalog ({source['id']})",
        "",
        "The harvester copies the guild's content verbatim and flags what looks off.",
        "Nothing here was 'fixed' — these are questions for the guild to rule on.",
        "",
    ]
    if anomalies:
        lines += [f"{i}. {format_anomaly(a)}" for i, a in enumerate(anomalies, 1)]
    else:
        lines.append("No anomalies found.")
    lines.append("")
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines))


def harvest(source):
    adapter = ADAPTERS.get(source["adapter"])
    if adapter is None:
        raise SystemExit(f"unknown adapter: {source['adapter']} (have: {', '.join(ADAPTERS)})")
    catalog, anomalies, prov = adapter(source)

    out = os.path.normpath(os.path.join(HERE, source["output"]))
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, "w", encoding="utf-8", newline="\n") as f:
        json.dump(catalog, f, indent=2, ensure_ascii=False)
        f.write("\n")

    stem = out[: -len(".json")] if out.endswith(".json") else out
    anomalies_path = stem + "-anomalies.md"
    write_anomalies(anomalies_path, source, anomalies)

    # the provenance sidecar: the leader-facing record of what the adapter did.
    # It must agree with the catalog — a mismatch is a harvester bug, not an anomaly.
    sidecar = prov.finish(anomalies, len(catalog["quests"]))
    catalog_ids = {q["quest_id"] for q in catalog["quests"]}
    prov_ids = [r["quest_id"] for t in sidecar["tabs"] for r in t["rows"]
                if r["outcome"] == "quest"]
    stray = [qid for qid in prov_ids if qid not in catalog_ids]
    if stray or len(prov_ids) != len(catalog["quests"]):
        raise SystemExit(
            f"[{source['id']}] provenance disagrees with the catalog: "
            f"{len(prov_ids)} recorded quest rows vs {len(catalog['quests'])} quests"
            + (f"; unknown quest_ids {stray[:5]}" if stray else "")
        )
    provenance_path = stem + "-provenance.json"
    with open(provenance_path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(sidecar, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"[{source['id']}] {len(catalog['quests'])} quest(s) -> {os.path.relpath(out, HERE)}")
    print(f"[{source['id']}] {len(anomalies)} anomaly(ies) -> {os.path.relpath(anomalies_path, HERE)}")
    print(f"[{source['id']}] provenance ({sidecar['counts']['rows_seen']} row(s)) -> {os.path.relpath(provenance_path, HERE)}")
    return catalog


def main():
    with open(os.path.join(HERE, "sources.json"), encoding="utf-8-sig") as f:
        config = json.load(f)

    wanted = sys.argv[1] if len(sys.argv) > 1 else None
    ran = 0
    for source in config["sources"]:
        if wanted is not None and source["id"] != wanted:
            continue
        if wanted is None and not source.get("enabled", True):
            print(f"[{source['id']}] disabled — skipped ({source.get('note', '')})")
            continue
        harvest(source)
        ran += 1

    if ran == 0:
        raise SystemExit(f"no source matched {wanted!r} — check sources.json")


if __name__ == "__main__":
    main()
